import openpyxl as xl

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']
print(cell.value)
print(sheet.max_row) # To find how many rows in the spreadsheet

for row in range(2, sheet.max_row + 1): # Ignore header row but include last row
    print(row) # count how many rows
    cell = sheet.cell(row, 3)
    print(cell.value) # print value in all rows in column 3, excluding header row
    corrected_price = cell.value * 0.9

    corrected_price_cell = sheet.cell(row, 4) # Setting column 4 as the destination cells
    corrected_price_cell.value = corrected_price # Setting the values to the right ones

wb.save('tranactions2.xlsx') # Save updates prices to the correct cells in a new workbook file

from openpyxl.chart import BarChart, Reference

values = Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4) #Making a chart using data from row 2-4

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2') # Chart top left corner at cell e2

wb.save('tranactions2.xlsx')